from globalvars import *
import threading
import time
from calculation import *
import pandas as pd
import os
 
price_keys=('-INPUTPRICESS304SHEET-','-INPUTPRICENICKELSHEET-','-INPUTPRICESS316SHEET-','-INPUTPRICEMONELSHEET-', '-INPUTPRICEINCONELSHEET-','-INPUTPRICEHASTELLOYSHEET-','-INPUTPRICEALUMINIUMSHEET-','-INPUTPRICETITANIUMSHEET-','-INPUTPRICECOPERSHEET-',
           '-INPUTPRICECSSHEET-','-INPUTPRICEPTFESHEET-', '-INPUTPRICEPVCSHEET-', '-INPUTPRICESILVERSHEET-', '-INPUTPRICESHAFT304-','-INPUTPRICESHAFT316-','-INPUTPRICESHAFTMONEL-','-INPUTPRICESHAFTINCONEL-',
           '-INPUTPRICESHAFTHASTELLOY-','-INPUTPRICESHAFTALUMINIUM-','-INPUTPRICESHAFTCOPPER-','-INPUTPRICESHAFTTITANIUM-', '-INPUTPRICESHAFTCS-','-INPUTPRICESHAFTSILVER-', '-INPUTPRICESTESTING-',
           '-INPUTPRICESTAG-', '-INPUTPRICEMILLING-','-INPUTPRICEMACHINERY-', '-INPUTPRICEWELDING-','-INPUTPRICEPAINT-', '-INPUTPRICESWATERJET-','-INPUTPRICEWIRECUT-','-INPUTPRICEBOXING-','-INPUTPRICESHIPPING-')
#make a list of materials
def mylister(filename = dbfilename,sheetname='size',colname='index'):
    
    sh = pd.read_excel(dbfilename, sheet_name =sheetname.lower())
    mr=[]
    for s in sh[str(colname).lower()]:
        mr.append(str(s))
    return mr

def long_operation_thread(filename = dbfilename,sheetname='size',colname='index'):
    """
    A worker thread that communicates with the GUI through a global message variable
    This thread can block for as long as it wants and the GUI will not be affected
    :param seconds: (int) How long to sleep, the ultimate blocking call
    """
    
    thread = threading.Thread(target=mylister, args=(filename ,sheetname,colname), daemon=True)
    thread.start()
    
    

def structsheetfromreadedexcel__(sn):
    dbfilename='mainDB_.xlsx'

    import pandas as pd

    hsheet = pd.read_excel(dbfilename, sheet_name =str(sn.lower()), index_col=None, header=0)
    df = pd.DataFrame(hsheet)    
    data=[]
    r=[]
    s=[]
    for row in df.itertuples():
        r = row[1:]  # Skip the index
        data.append(r)
    return data    

def structsheetfromreadedexcel(snsheet):
    dsheet = gdata_[snsheet]
    df = pd.DataFrame(dsheet)
    data=[]
    r=[]
    s=[]
    for row in df.itertuples():
        r = row[1:]  # Skip the index
        data.append(r)
    return data        

    
def getoverqtyrupturefortest(rqty=2):
    data = structsheetfromreadedexcel('test')
    nqty=2
    for d in data:
        az = float(d[1])
        ta = float(d[2])
        if float(d[2])>=rqty and float(d[1])<=rqty:
            return int(max(float(d[3]),(rqty*float(d[4]))))
            
    return int(rqty*0.033)



def readresorcemakesource(filename):
    global gdata_
    
    
    global Global_Materials
    Global_Materials= structsheetfromreadedexcel__('material')
    
    global Global_Reverse
    Global_Reverse= structsheetfromreadedexcel__('reverse')
    
    global Global_Forward
    Global_Forward= structsheetfromreadedexcel__('forward')
    
    global Global_Flat
    Global_Flat= structsheetfromreadedexcel__('flat')
   
    global Global_Size
    Global_Size= structsheetfromreadedexcel__('size')
    
    global Global_Mto
    Global_Mto= structsheetfromreadedexcel__('mto')
    
    global Global_testqty
    Global_testqty= structsheetfromreadedexcel__('test')
    
    
    gdata_={'material':Global_Materials,'test':Global_testqty,'mto':Global_Mto,'size':Global_Size,'reverse':Global_Reverse,'forward':Global_Forward,'flat':Global_Flat}
    return gdata_
    
    
def findmaterialifpredefined(mat='s316'):
    data = structsheetfromreadedexcel('material')
    m = str(mat).lower()
    if len(m)>3:
        m=m[0:3]
        
    
    for d in data:
        s=str(d[2]).lower()
        if len(s)>2:
            s=s[0:3]
        
        if s==m:
            return d
    ndd=[]
    ndd.append(0)
    ndd.append(str(mat))
    ndd.append(mat)
    ndd.append(mat)
    ndd.append(5)
    ndd.append(5000000)
    return ndd




def getmaterialnamefa(mat='s316'):
    mf = findmaterialifpredefined(mat)
    return str(mf[3])

def getmaterialnamepriceunit(mat='s316'):
    mf = findmaterialifpredefined(mat)
    return str(mf[5])

   
   
def getrupturelayers(rtype='reverse'):

    
    data = structsheetfromreadedexcel(rtype)

    dd=[]
    r=[]
    mr=[]
    m=[]
    for d in data:
        r=[]
        mr=[]
        for i in range(5):
            m=[]
            if not pd.isna(d[3+i*2]):
                m=[]
                m.append(d[3+i*2])
                m.append(d[3+1+i*2])
                mr.append(m)
        r.append(mr)        
        m=[]
        m.append(d[13])
        m.append(d[14])
        m.append(d[15])
        r.append(m)
        m=[]
        m.append(d[2])
        m.append(d[16])
        m.append(d[17])        
        r.append(m)
        dd.append(r)
    return dd 


        
def findsizeequalityvalue(rtype='reverse',rsize=2):
    data = getrupturelayers(rtype)
    sized=[]
    for test in data:
        tc = test[2]
        if tc[0]==rsize:
            sized.append(test)
    return sized
            
 
def sortingbyrburstpressure(rtype='reverse',rsize=2,rbp=5):
    datas= findsizeequalityvalue(rtype,rsize)
    data=[]
    for rds in datas:
        rd = rds[1]
        rds.append(abs(rbp-float(rd[2])))
        data.append(rds)
    sdata=     sorted(data,key=lambda x:x[3])
    return sdata
     
def findneardesignedruptureperv(rtype='reverse',rsize=2,rbp=5):
    data = sortingbyrburstpressure(rtype,rsize,rbp)
    if len(data)>0:
        return data[0]
    else:
        rsize=rsize+2
        data = sortingbyrburstpressure(rtype,rsize,rbp)
        return data[0]
    
        
def getdimensionbysizetype(element='rupture',etype='reverse',esize=2):
    data = structsheetfromreadedexcel('size')
    element=element.lower()
    etype=etype.lower()
    esize=int(float(esize)*10)
    r=[]
    for d in data:
        if d[0]==element:
            
            if  etype=='' or d[1]== etype:
                if (esize == 0) or (int(d[2]*10) == esize):
                    for i in range(4,len(d)):
                        if d[i]!=0:
                            r.append(float(d[i]))
                            
                
    return r

                    
def getreadmtoitemorop(element='rupture'):
    data = structsheetfromreadedexcel('mto')
    r=[]
    it=[]
    rd=[]
    for d in data:
        if d[0]==element:
            r=[]
            it=[]
            it.append(d[2])
            it.append(d[3])
            r.append(it)
            it=[]
            it.append(d[4])
            it.append(d[5])
            r.append(it)
            it=[]
            it.append(d[6])
            it.append(d[7])
            r.append(it)    
            rd.append(r)
    return rd   

def getreadmtoheader(l='Fa'):
    data = structsheetfromreadedexcel('mto')
    r=[]
    it=[]
    for d in data:
        if d[0]=='mtoheader':
            
            if l=='Fa':
            
                it.append(d[3])
            else:
                it.append(d[2])
            r.append(it)
            
    return it





def getmyfilename(pt='../my_data/',fn='1',rt='reverse',rs=4,rb=11,rq=10):
    filename=""
    if not os.path.exists(pt): 
          
        # if the demo_folder directory is not present  
        # then create it. 
        os.makedirs(pt)     
    if fn.find('__')>-1:
        filename = fn[0:fn.find('__')]
    if fn.find('.')>-1:
        filename = fn[0:fn.find('.')]
        fileexo = fn[fn.find('.'):len(fn)]
    else:
        fileexo = fileexepdef
        for f in fn :
            if f in ('0123456789'):
                filename = filename.__add__(f) 
    
                
    file_path_name = pt.__add__(filename).__add__('__type ').__add__(str(rt)).__add__('_size ').__add__(str(rs)).__add__('_BP ').__add__(str(rb)).__add__('_QTY ').__add__(str(rq)).__add__('.').__add__(fileexo)
    # Check if the file exists
    while os.path.exists(file_path_name):
        filename = str(int(filename)+1)
        file_path_name = pt.__add__(filename).__add__('__type ').__add__(str(rt)).__add__('_size ').__add__(str(rs)).__add__('_BP ').__add__(str(rb)).__add__('_QTY ').__add__(str(rq)).__add__('.').__add__(fileexo)
    return file_path_name

         
def excelmtofilestylerfarsi(fn='mto',shn='_mto',title='Project'):
    
 
    from openpyxl import load_workbook 
    from openpyxl.styles import Font, PatternFill, Alignment ,Border , Side
    from openpyxl.utils.dataframe import dataframe_to_rows 
    # Sample DataFrame to write to Excel 
    # Write DataFrame to Excel file 
    file_path = fn
    
    # Load the workbook and the sheet 
    wb = load_workbook(file_path) 
    
    sheet = wb[shn] 
    sheet.sheet_properties.rightToLeft = True 
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))      
    borderH = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick'))       
    # Set font for the rest of the sheet and wrap text 
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=8): 
        for cell in row: cell.font = Font(name="B Nazanin", size=14) 
        cell.alignment = Alignment(wrap_text=True) 
        cell.border = border
        # Wrap text in each cell 
        # Set header row (Row 2) 
    header_row = sheet[1] 
    for cell in header_row: 
        cell.font = Font(size=16,bold=True) 
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True) 
        cell.fill = PatternFill(start_color="90A4FA", end_color="90A4FA", fill_type="solid") 
        # Light gray background 
    # Set row background colors (odd/even rows) 
      
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=8), start=2):
        for cell in row: 
            if i % 2 == 0: 
                cell.fill = PatternFill(start_color="9BEDEF", end_color="9BEDEF", fill_type="solid")
                cell.border = border
                # Even row light gray 
            else: 
                cell.fill = PatternFill(start_color="CBF298", end_color="CBF298", fill_type="solid")
                cell.border = border
                # Odd row white 
    # Set sheet direction to right to left 
    sheet.sheet_properties.rightToLeft = True 
    
    # Resize columns to fit the content 
    colconter=0
    coladder = 2
    for col in sheet.columns:
        max_length = 0 
        column = col[0].column_letter 
        # Get the column name (e.g., 'A', 'B', etc.) 
        for cell in col: 
            try: 
                if len(str(cell.value)) > max_length: 
                    max_length = len(cell.value) 
            except: 
                pass 
            if colconter == 0:
                coladder=3
            elif colconter ==1:
                coladder=8
            elif colconter == 2:
                coladder = 25
            elif colconter == 3:
                coladder = 4
            elif colconter == 4:
                coladder = 4
            elif colconter>4:
                coladder = 10
                
               
            adjusted_width = (max_length + coladder) 
            sheet.column_dimensions[column].width = adjusted_width 
   
            
    
   

    # Select the active sheet
    sheet = wb.active

    # Insert an empty row at the top
    sheet.insert_rows(1)

    
    # Set title in the first row and merge cells 
    sheet.merge_cells('A1:H1')
    title_cell = sheet['A1'] 
    title_cell.value = title
    title_cell.font = Font(name="Arial",size=18, bold=True) 
    title_cell.alignment = Alignment(horizontal='center', vertical='center') 
    title_cell.fill = PatternFill(start_color="FFFF25", end_color="FFFF25", fill_type="solid") 
    title_cell.border = borderH
    #title_cell.border = borderH
    # Yellow background
    print("Excel file styled and saved with wrapped text and auto-resized columns.") 
    
    
    # Save the workbook 
    wb.save(file_path)    

  
def writedataframedstylishtofile(df,fn,title):
    df.to_excel(fn,sheet_name='_mto',index=False) 
    excelmtofilestylerfarsi(fn,'_mto',title)    
    return fn

            



def getdensityofmaterial(mat='s316'):
    mf = findmaterialifpredefined(mat)
    return str(mf[4])    

def getpriceofmaterialkg(mat):
    mf = findmaterialifpredefined(mat)
    return str(mf[5])  

def findpricefromconfigprice(mat='s316',tip='SHEET',val=[]):
    mat=mat.upper()
    tip=tip.upper()
    for key in keys_to_save:
        if str(key).find(mat)>-1 and tip in str(key):
            return val[key]
                
    return getpriceofmaterialkg(mat=mat)     


def getunitpriceformtoitem(mtoitem,itemnum):
    unit = mtoitem[itemnum]
    return unit[2]

def getreadpriceformtoitems(mtoitem=[],itemnumber=4,v=[]):
    mtp=[]

    pass

def getholdersizes(etype,esize):
    hdim=getdimensionbysizetype(element='holder',etype=etype,esize=esize)
    return hdim
    
def findingneareststandardshaftforholder(htype='reverse',hsize=4):
    hs= getholdersizes(htype,hsize)
    return hs[2]

def getprocesssideholderthickness(htype,hsize):
    hs= getholdersizes(htype,hsize)
    return hs[4]    
def getventsideholderthickness(htype,hsize):
    hs= getholdersizes(htype,hsize)
    return hs[5] 
def roundshaftlengthformachinery(l):
    
    l/=5
    
    l+=2
    l=5*int(l)
    return int(l)
